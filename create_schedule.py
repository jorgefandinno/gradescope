import argparse

parser = argparse.ArgumentParser(description="Updates the schedule to a new semester")
parser.add_argument("old_schedule", type = argparse.FileType('r'), help="Path to the old schedule")
args = parser.parse_args()
args.old_schedule.close()

##########################################################################################################
##########################################################################################################
##########################################################################################################

from datetime import datetime, timedelta

YEAR = 2024
SEMESTER = "spring"

FIST_DAY_OF_CLASSES = datetime(YEAR, 1, 22)
LAST_DAY_OF_CLASSES = datetime(YEAR, 5, 11)
WEEK_DAY_OF_CLASSES = 0 # 0 is MW, 1 is TR

LIST_OF_HOLIDAYS = [
    (datetime(YEAR, 3, 10), datetime(YEAR, 3, 17)),
]

EXAMS = {
    datetime(YEAR, 3, 6) : "Midterm 1",
    datetime(YEAR, 4, 17) : "Midterm 2",
    datetime(YEAR, 5,  8) : "Midterm 3",
}


##########################################################################################################
##########################################################################################################
##########################################################################################################

import re
from pathlib import Path
from pprint import pprint

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import NamedStyle

list_of_holidays = [a if isinstance(a, tuple) else (a,a) for a in LIST_OF_HOLIDAYS]
holy_days = set(start + timedelta(days=day) for start, end in LIST_OF_HOLIDAYS for day in range((end - start).days+1))

def new_schedule_name(old_schedule_name):
    year = re.findall(r"\d{4}", old_schedule_name)
    if len(year) == 0:
        year = re.findall(r"\d{2}", old_schedule_name)
    if len(year) >= 1:
        return old_schedule_name.replace(year[0], str(YEAR)).replace("spring", SEMESTER).replace("fall", SEMESTER)
    else:
        return old_schedule_name.replace("spring", SEMESTER).replace("fall", SEMESTER) + f"_{YEAR}"

old_schedule_file = Path(args.old_schedule.name)
new_schedule_file = old_schedule_file.parent.joinpath(new_schedule_name(old_schedule_file.stem)+old_schedule_file.suffix)

# print(f"Old schedule: {old_schedule_file}")
# print(f"New schedule: {new_schedule_file}")

def read_old_schedule(excel_file_path):
    topics = []
    workbook = load_workbook(excel_file_path)
    sheet = workbook[workbook.sheetnames[0]]
    topic_cells = []
    topics = []
    assigments = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        topic_cells.append(row[2])
        topic_cells.append(row[5])
        if row[0].value is not None:
            assigments.append((row[6].value, row[7].value, row[8].value))

    for cell in topic_cells:
        content = cell.value
        fill_color = cell.font.color.rgb
        if content is not None and not content.strip().lower().startswith("recap") and fill_color != "FFFF0000":
            topics.append(content)

    return topics, assigments


def _get_topic(list_of_topics, date, topic_index, second_date=None):
    if date in holy_days:
        topic = "Holiday"
    elif date in EXAMS:
        topic = EXAMS[date]
    elif second_date is not None and second_date in EXAMS:
        topic = "Recap"
    elif topic_index < len(list_of_topics):
        topic = list_of_topics[topic_index]
        topic_index += 1
    else:
        topic = ""
    return topic, topic_index


topics, assigments = read_old_schedule(old_schedule_file)

date = FIST_DAY_OF_CLASSES
topic_index = 0
assigments_index = 0
data = []
while date <= LAST_DAY_OF_CLASSES:
    day1  = date + timedelta(WEEK_DAY_OF_CLASSES)
    day2 = date + timedelta(WEEK_DAY_OF_CLASSES+2)
    topic1, topic_index = _get_topic(topics, day1, topic_index, day2)
    topic2, topic_index = _get_topic(topics, day2, topic_index)
    row = [date, day1, topic1, "", day2, topic2] + list(assigments[assigments_index])
    data.append(row)
    date += timedelta(days=7)
    assigments_index += 1

data = pd.DataFrame(data, columns=["Week", "Date1", "Topic1", "", "Date2", "Topic2", "Projects", "HW", "Readings"])
# data.to_excel("schedule_pl_2024.xlsx", index=False)

with pd.ExcelWriter(new_schedule_file, engine='openpyxl') as writer:
    data.to_excel(writer, sheet_name='Sheet1', index=False)

    # Get the workbook and the sheet
    workbook = writer.book
    sheet = writer.sheets['Sheet1']

    sheet.column_dimensions['A'].width = 15  # Width of column 'A'
    sheet.column_dimensions['B'].width = 15  # Width of column 'B'
    sheet.column_dimensions['C'].width = 50  # Width of column 'C'
    sheet.column_dimensions['D'].width = 10  # Width of column 'D'
    sheet.column_dimensions['E'].width = 15  # Width of column 'E'
    sheet.column_dimensions['F'].width = 50  # Width of column 'F'
    sheet.column_dimensions['G'].width = 15  # Width of column 'G'
    sheet.column_dimensions['H'].width = 10  # Width of column 'H'
    sheet.column_dimensions['I'].width = 10  # Width of column 'I'

    # Define date formatting style
    date_format = NamedStyle(name='date_format', number_format='MM/DD/YYYY')
    red_format  = NamedStyle(name='red_format', font=Font(color='FF0000'))
    workbook.add_named_style(date_format)
    for row in range(2, len(data)+2):
        sheet[f'A{row}'].style = date_format
        sheet[f'B{row}'].style = date_format
        sheet[f'E{row}'].style = date_format
        cell_content = sheet[f'C{row}'].value.strip()
        if cell_content.startswith("Holiday") or cell_content.startswith("Midterm"):
            sheet[f'C{row}'].style = red_format
        cell_content = sheet[f'F{row}'].value.strip()
        if cell_content.startswith("Holiday") or cell_content.startswith("Midterm"):
            sheet[f'F{row}'].style = red_format
            

